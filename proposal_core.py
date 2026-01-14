import io
import re
import unicodedata
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

# ----------------------------------------------------
# 공통 유틸: 문자열 정규화 / 유전자(2-1~2-4) 항목 판별
# ----------------------------------------------------
DASH_CHARS = "–—−‐‑‒﹣－ㅡ"

def normalize_key(text):
    """공백/특수 대시/유사문자 등을 표준화하여 비교 안정성 확보"""
    if text is None:
        return ""
    s = unicodedata.normalize("NFKC", str(text))

    # 대시처럼 보이는 문자들 통일
    for ch in DASH_CHARS:
        s = s.replace(ch, "-")

    # 자주 섞이는 유사문자(키보드 입력 실수 등) 보정
    s = s.replace("ㅡ", "-")

    # 공백 제거
    s = re.sub(r"\s+", "", s)

    # 맨 앞 글머리표/특수문자 제거(예: '•2-1', '(2-1)')
    s = re.sub(r"^[^0-9]*", "", s)
    return s


def is_gene_block_item(name):
    """'2-1~2-4'로 시작하고 '유전자'가 포함된 항목인지 판별"""
    key = normalize_key(name)
    return ("유전자" in key) and bool(re.match(r"^2[-._]*([1-4])", key))



def extract_gene_num(name):
    """유전자 항목의 번호(1~4)를 추출. 실패 시 None."""
    key = normalize_key(name)
    if "유전자" not in key:
        return None
    m = re.match(r"^2[-._]*([1-4])", key)
    return int(m.group(1)) if m else None


def find_gene_block_indices(items):
    """EQUIP 항목 중 '유전자 4행(2-1~2-4)'에 해당하는 행 인덱스를 찾는다.
    - 1순위: 2-1~2-4 번호가 모두 잡히는 경우
    - 2순위: '유전자'가 포함된 행 중 연속된 4개
    - 3순위: '유전자' 포함 행이 4개 이상이면 앞 4개
    """
    num_to_idx = {}
    for i, it in enumerate(items):
        num = extract_gene_num(it.get("name", ""))
        if num is not None:
            num_to_idx[num] = i

    if all(n in num_to_idx for n in (1, 2, 3, 4)):
        return [num_to_idx[n] for n in (1, 2, 3, 4)]

    gene_idxs = [i for i, it in enumerate(items) if "유전자" in normalize_key(it.get("name", ""))]
    if len(gene_idxs) < 4:
        return []

    # 연속된 4개 찾기
    for start in range(len(gene_idxs) - 3):
        block = gene_idxs[start:start + 4]
        if block[0] + 1 == block[1] and block[1] + 1 == block[2] and block[2] + 1 == block[3]:
            return block

    # 연속이 아니어도 4개 이상이면 앞 4개 사용(최후의 안전장치)
    return gene_idxs[:4]


def apply_gene_block_fix_30only(equip_items, plans):
    """유전자(2-1~2-4) 블록 보정:
    - 30만원 플랜 열에서만 4개 행 모두 '선택 1'로 채움(병합 가능해짐)
    - 병합 필터용 플래그(_force_merge_gene) 설정
    """
    idxs = find_gene_block_indices(equip_items)
    if not idxs:
        return False

    # 병합 플래그
    for i in idxs:
        equip_items[i]["_force_merge_gene"] = True

    # 30만원 플랜에서만 값 강제
    for p_idx, plan in enumerate(plans):
        if plan.get("sort_key", 0) == 30:
            for i in idxs:
                equip_items[i]["values"][p_idx] = "선택 1"

    return True


def scan_default_counts(ws, col_idx, start_row):
    """엑셀에서 '선택 N'을 스캔하여 기본값 추출"""
    counts = {'a': 0, 'b': 0, 'c': 0}
    max_scan = min(start_row + 150, ws.max_row)
    current_cat = ""

    for r in range(start_row + 1, max_scan + 1):
        c1_val = ws.cell(row=r, column=1).value
        c_target_val = ws.cell(row=r, column=col_idx).value
        
        cell_group = str(c1_val).strip() if c1_val else ""
        cell_val = str(c_target_val).strip() if c_target_val else ""

        if "A그룹" in cell_group: current_cat = "a"
        elif "B그룹" in cell_group: current_cat = "b"
        elif "C그룹" in cell_group: current_cat = "c"

        if current_cat in ['a', 'b', 'c'] and "선택" in cell_val:
            nums = re.findall(r'\d+', cell_val)
            if nums:
                val = int(nums[0])
                if val > counts[current_cat]:
                    counts[current_cat] = val
    return counts

def load_price_options(excel_path):
    """엑셀 헤더 분석 및 기본 선택값 로드"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    header_row_idx = None
    price_cols = []

    for row in sheet.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value and "만원" in str(cell.value):
                header_row_idx = cell.row
                break
        if header_row_idx: break
    
    if not header_row_idx:
        return None, []

    row_cells = list(sheet.rows)[header_row_idx - 1]
    excluded = ["10만원", "15만원"]
    
    manual_defaults = {
        25: {'a': 3, 'b': 0, 'c': 0}, 30: {'a': 3, 'b': 0, 'c': 0},
        35: {'a': 4, 'b': 0, 'c': 0}, 40: {'a': 5, 'b': 0, 'c': 0},
        45: {'a': 4, 'b': 1, 'c': 0}, 50: {'a': 5, 'b': 1, 'c': 0},
        60: {'a': 3, 'b': 1, 'c': 1}, 70: {'a': 5, 'b': 1, 'c': 1},
        80: {'a': 5, 'b': 2, 'c': 1}, 90: {'a': 5, 'b': 3, 'c': 1},
        100: {'a': 3, 'b': 3, 'c': 2}
    }

    for idx, cell in enumerate(row_cells):
        val = str(cell.value).strip() if cell.value else ""
        if "만원" in val and not any(e in val for e in excluded):
            col_idx = idx + 1
            try: price_num = int(re.sub(r'[^0-9]', '', val))
            except: price_num = 0

            if price_num in manual_defaults:
                defaults = manual_defaults[price_num]
            else:
                defaults = scan_default_counts(sheet, col_idx, header_row_idx)
            
            price_cols.append({
                "price_txt": val,
                "col_idx": col_idx,
                "defaults": defaults,
                "sort_key": price_num
            })
    
    wb.close()
    price_cols.sort(key=lambda x: x['sort_key'])
    return header_row_idx, price_cols

def parse_data_from_excel(excel_path, header_row, plans):
    """엑셀 데이터 파싱"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    parsed_data = {"A": [], "B": [], "C": [], "EQUIP": [], "COMMON_BLOOD": []}
    summary_info = []

    for p in plans:
        summary_info.append({
            "name": p["name"],
            "a": p.get("a_rule", "-"),
            "b": p.get("b_rule", "-"),
            "c": p.get("c_rule", "-")
        })

    fill_cache = {i: {"A": None, "B": None, "C": None} for i in range(len(plans))}
    current_main_cat = ""
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or len(row) < 2: continue
        col0 = str(row[0]).strip() if row[0] else ""
        col1 = str(row[1]).strip() if row[1] else ""
        col0_clean = col0.replace(" ", "")

        if "A그룹" in col0_clean: current_main_cat = "A"
        elif "B그룹" in col0_clean: current_main_cat = "B"
        elif "C그룹" in col0_clean: current_main_cat = "C"
        elif "장비검사" in col0_clean or "소화기검사" in col0_clean: current_main_cat = "EQUIP"
        elif "혈액" in col0_clean and "소변" in col0_clean: current_main_cat = "COMMON"

        if not col1 or col1 in ["검진항목", "내용"]: continue

        item_name = col1
        item_desc = str(row[2]).strip() if row[2] else ""
        sub_cat = col0 if current_main_cat == "EQUIP" and col0 else ""

        # [수정됨] 에피클락 제거 로직
        if "에피클락" in item_name:
            continue
        # [수정] 유전자(2-1~2-4) 항목 판별: 공백/특수대시 차이를 흡수하여 인식 안정화
        is_target_gene = is_gene_block_item(item_name)
        row_vals = []
        for idx, plan in enumerate(plans):
            col_idx = plan["col_idx"] - 1
            val = ""
            if col_idx < len(row):
                val = str(row[col_idx]).strip() if row[col_idx] else ""

            # [수정됨] 유전자 항목인 경우 가격 조건에 따라 값 강제 변경
            if is_target_gene:
                plan_price = plan.get('sort_key', 0)
                if plan_price == 30:
                    val = "선택 1"
                # 30만원이 아니면 엑셀 원래 값 유지
            else:
                # 일반 항목 처리 로직 (선택 N 캐싱 등)
                if current_main_cat in ["A", "B", "C"]:
                    cache = fill_cache[idx]
                    if "선택" in val: cache[current_main_cat] = val
                    elif val == "" and cache[current_main_cat]: val = cache[current_main_cat]
                    elif val != "": cache[current_main_cat] = None
                    
                    if "선택" in val:
                        rule = ""
                        if current_main_cat == "A": rule = plan.get('a_rule', '')
                        elif current_main_cat == "B": rule = plan.get('b_rule', '')
                        elif current_main_cat == "C": rule = plan.get('c_rule', '')
                        if rule:
                            val = "" if rule == "-" else rule

            if "미선택" in val: val = ""
            row_vals.append(val)

        entry = {"category": sub_cat, "name": item_name, "desc": item_desc, "values": row_vals}
        
        if current_main_cat == "A": parsed_data["A"].append(entry)
        elif current_main_cat == "B": parsed_data["B"].append(entry)
        elif current_main_cat == "C": parsed_data["C"].append(entry)
        elif current_main_cat == "EQUIP": parsed_data["EQUIP"].append(entry)
        elif current_main_cat == "COMMON": parsed_data["COMMON_BLOOD"].append(entry)

    # [추가] 유전자(2-1~2-4) 블록 보정: 30만원 플랜에서만 선택 1 채우고 병합 플래그 부여
    apply_gene_block_fix_30only(parsed_data.get("EQUIP", []), plans)

    wb.close()
    return parsed_data, summary_info

def render_html_string(plans, data, summary, info):
    """HTML 생성"""
    today_date = datetime.now().strftime("%Y년 %m월 %d일")
    company = info.get('company', '')
    manager = info.get('name', '담당자')
    proposal_title = f"2026 {company} 임직원 건강검진 제안서" if company else "2026 기업 임직원 건강검진 제안서"

    def normalize_text(text):
        return re.sub(r'(선택)\s*(\d+)', r'\1 \2', str(text))
    
    def get_val_display(val):
        if not val or val in ['X', 'x', '-', '미선택']: return ""
        if val in ['O', 'o', '○'] or "기본" in str(val): return "O"
        if "선택" in val: return normalize_text(val)
        return val

    def render_table_html(title, item_list, show_sub=False, footer=None, merge=True, merge_filter=None):
        if not item_list: return ""
        
        grid = []
        for item in item_list:
            row = [get_val_display(v) for v in item['values']]
            grid.append(row)
        
        rows_cnt = len(grid)
        cols_cnt = len(plans)
        rowspan_map = [[1] * cols_cnt for _ in range(rows_cnt)]
        skip_map = [[False] * cols_cnt for _ in range(rows_cnt)]

        can_merge_row = [merge_filter(it) for it in item_list] if merge_filter else [True] * rows_cnt

        if merge:
            for c in range(cols_cnt):
                for r in range(rows_cnt):
                    if skip_map[r][c]: continue
                    val = grid[r][c]
                    if val != "" and can_merge_row[r]:
                        span = 1
                        for k in range(r + 1, rows_cnt):
                            # [중요] 값이 같으면 병합 (선택 1 == 선택 1 이면 병합됨)
                            if can_merge_row[k] and grid[k][c] == val:
                                span += 1; skip_map[k][c] = True
                            else: break
                        rowspan_map[r][c] = span
        
        html_rows = ""
        for r in range(rows_cnt):
            item = item_list[r]
            sub_tag = f"<span class='cat-tag'>[{item['category']}]</span> " if show_sub and item['category'] else ""
            
            name_style = ""
            if "스마트암검사" in item['name']:
                name_style = " style='white-space:nowrap; letter-spacing:-1.5px;'"
            
            tr_style = ""
            if item['name'] == "우대수가":
                tr_style = " style='background-color:#F0F2F5; font-weight:bold;'"

            row_str = f"<tr{tr_style}><td class='item-name-cell'><div{name_style}>{sub_tag}{item['name']}</div></td>"
            
            for c in range(cols_cnt):
                if skip_map[r][c]: continue
                val = grid[r][c]
                span = rowspan_map[r][c]
                cls = "text-center"
                if val == "O": cls += " text-bold"
                elif "선택" in str(val): cls += " text-navy text-bold"
                attr = f' rowspan="{span}"' if span > 1 else ""
                row_str += f'<td{attr} class="{cls}">{val}</td>'
            row_str += "</tr>"
            html_rows += row_str

        header_cols = "".join([f"<th>{p['name']}</th>" for p in plans])
        footer_div = f"<div class='table-footer'>{footer}</div>" if footer else ""
        
        return f"""
        <div class="section">
            <div class="sec-title">{title}</div>
            <table>
                <thead><tr><th style="width:28%">검사 항목</th>{header_cols}</tr></thead>
                <tbody>{html_rows}</tbody>
            </table>
            {footer_div}
        </div>
        """

    a_vals = [s.get('a', '-') for s in summary]
    b_vals = [s.get('b', '-') for s in summary]
    c_vals = [s.get('c', '-') for s in summary]
    
    def make_sum_row(title, vals):
        tds = "".join([f"<td class='text-center'>{v}</td>" for v in vals])
        return f"<tr><td class='summary-header'>{title}</td>{tds}</tr>"
    
    sum_rows_html = make_sum_row("A그룹", a_vals) + make_sum_row("B그룹", b_vals) + make_sum_row("C그룹", c_vals)
    sum_headers = "".join([f"<th>{p['name']}</th>" for p in plans])

    css = """
    @import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard/dist/web/static/pretendard.css');
    body { font-family: 'Pretendard', sans-serif; background: #fff; margin: 0; padding: 20px; color: #333; font-size: 11px; }
    .page { width: 210mm; margin: 0 auto; background: white; padding: 15px 40px; box-sizing: border-box; position: relative; }
    
    /* Cover Page CSS */
    .cover-container { 
        width: 100%; height: 280mm; 
        display: flex; flex-direction: column; justify-content: space-between; 
        padding: 40px 20px; box-sizing: border-box;
        border: 1px solid #ddd;
    }
    .cover-top { text-align: right; border-bottom: 2px solid #1a253a; padding-bottom: 10px; margin-bottom: 20px; }
    .cover-top-title { font-size: 14px; font-weight: bold; color: #555; letter-spacing: 1px; }
    .cover-top-date { font-size: 12px; color: #7f8c8d; margin-top: 5px; }
    .cover-middle { text-align: center; margin-top: 60px; margin-bottom: 60px; }
    .cover-main-title { font-size: 26px; font-weight: bold; color: #333; letter-spacing: 2px; margin-bottom: 40px; }
    .cover-client-name { font-size: 42px; font-weight: 900; color: #1a253a; line-height: 1.3; border-bottom: 1px solid #ccc; display: inline-block; padding-bottom: 10px; margin-bottom: 10px; }
    .cover-honorific { font-size: 24px; font-weight: 500; color: #555; margin-left: 10px; }
    .cover-bottom { text-align: center; margin-bottom: 30px; }
    .cover-submit-box { display: inline-block; text-align: left; border-top: 2px solid #1a253a; padding-top: 20px; margin-top: 50px; }
    .cover-submit-row { margin-bottom: 12px; font-size: 15px; }
    .cover-submit-label { display: inline-block; width: 80px; font-weight: bold; color: #555; }
    .cover-submit-val { font-weight: bold; color: #1a253a; font-size: 16px; }

    /* Content CSS */
    .hospital-brand { font-size: 26px; font-weight: 900; color: #1a253a; letter-spacing: -1px; }
    .hospital-sub { font-size: 16px; color: #555; margin-top: 5px; font-weight: bold; }
    .contact-card { background-color: #f8f9fa; border: 2px solid #2c3e50; border-radius: 8px; padding: 10px 15px; text-align: right; box-shadow: 2px 2px 8px rgba(0,0,0,0.05); min-width: 200px; float: right; }
    header { display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 15px; }
    .header-divider { border-bottom: 2px solid #2c3e50; margin-bottom: 15px; clear: both; }
    .section { margin-bottom: 25px; page-break-inside: avoid; }
    .sec-title { font-size: 15px; font-weight: 800; color: #2c3e50; margin-bottom: 8px; padding-left: 8px; border-left: 4px solid #2c3e50; }
    table { width: 100%; border-collapse: collapse; table-layout: fixed; font-size: 11px; border-top: 2px solid #2c3e50; }
    th { background: #f0f2f5; color: #2c3e50; padding: 8px; border: 1px solid #bdc3c7; font-weight: bold; }
    td { padding: 6px; border: 1px solid #bdc3c7; vertical-align: middle; word-break: keep-all; height: 24px; }
    .summary-table th { background: #34495e; color: white; border-color: #2c3e50; }
    .summary-header { background: #f8f9fa; font-weight: bold; color: #2c3e50; padding-left: 15px; text-align: left; }
    .text-center { text-align: center; }
    .text-bold { font-weight: bold; }
    .text-navy { color: #2c3e50; }
    .item-name-cell { text-align:left; padding-left:10px; width: 28%; font-weight: 600; }
    .cat-tag { color: #7f8c8d; font-size: 10px; margin-right:3px; }
    .table-footer { font-size: 11px; color: #2c3e50; text-align: right; margin-top: 5px; font-weight: bold; }
    .guide-box { background-color: #fff; border: 2px solid #2c3e50; padding: 15px; margin-bottom: 15px; font-size: 11px; line-height: 1.6; color: #333; }
    .guide-title { font-weight: 800; font-size: 14px; margin-bottom: 10px; display:block; color: #2c3e50; border-bottom: 1px solid #ddd; padding-bottom: 5px; }
    .highlight-text { font-weight: bold; color: #1a253a; }
    .important-note { color: #c0392b; font-weight: bold; }
    .program-grid { display: flex; flex-direction: column; gap: 6px; margin-bottom: 20px; border: 1px solid #ccc; padding: 6px; background: #fff; }
    .grid-row { display: flex; gap: 6px; }
    .grid-col { display: flex; flex-direction: column; gap: 6px; }
    .grid-box { border: 1px solid #95a5a6; background: white; }
    .grid-header { background: #34495e; color: white; padding: 6px 10px; font-weight: bold; font-size: 12px; text-align: center; }
    .grid-content { padding: 10px; font-size: 11px; line-height: 1.5; color: #333; }
    .grid-content-list { display: grid; grid-template-columns: 1fr 1fr; gap: 2px 10px; padding: 8px 10px; font-size: 11px; font-weight: 500; color: #444; }
    .grid-sub-header { background: #ecf0f1; color: #2c3e50; padding: 4px 10px; font-weight: bold; font-size: 11px; border-bottom: 1px solid #ddd; }
    .header-common { background: #2c3e50; font-size: 13px; text-align: left; padding-left: 15px; }
    .header-a { background: #566573; }
    .header-b { background: #7f8c8d; }
    .header-c { background: #2c3e50; }
    @media print { .page { break-after: page; } .no-print { display: none; } }
    """

    head = f"""
    <!DOCTYPE html>
    <html lang="ko">
    <head>
        <meta charset="UTF-8">
        <style>{css}</style>
    </head>
    <body>
    """

    cover_html = f"""
        <div class="page">
            <div class="cover-container">
                <div class="cover-top">
                    <div class="cover-top-title">2026 기업 임직원 건강검진 제안서</div>
                    <div class="cover-top-date">{datetime.now().strftime('%Y-%m-%d')}</div>
                </div>
                <div class="cover-middle">
                    <div class="cover-main-title">HEALTH CHECK-UP PROPOSAL</div>
                    <div>
                        <span class="cover-client-name">{company}</span>
                        <span class="cover-honorific">귀중</span>
                    </div>
                </div>
                <div class="cover-bottom">
                    <div class="cover-submit-box">
                        <div class="cover-submit-row">
                            <span class="cover-submit-label">제 출 처</span>
                            <span class="cover-submit-val">뉴고려병원 검진사업부</span>
                        </div>
                        <div class="cover-submit-row">
                            <span class="cover-submit-label">담 당 자</span>
                            <span class="cover-submit-val">{manager} 팀장</span>
                        </div>
                        <div class="cover-submit-row">
                            <span class="cover-submit-label">문 의</span>
                            <span class="cover-submit-val">{info.get('phone','')}</span>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        <div style="page-break-after: always;"></div>
        <div class="page">
    """

    header_content = f"""
            <header>
                <div>
                    <div class="hospital-brand">뉴고려병원</div>
                    <div class="hospital-sub">{proposal_title}</div>
                    <div style="font-size:11px; color:#666; margin-top:4px;">제안일자: {today_date}</div>
                </div>
                <div class="contact-card">
                    <div style="font-size: 10px; color: #7f8c8d; font-weight: bold;">PROPOSAL CONTACT</div>
                    <div style="font-size: 14px; font-weight: 800; color: #2c3e50;">{info.get('name','')} 팀장</div>
                    <div style="font-size: 11px; font-weight: 600;">📞 {info.get('phone','')}</div>
                    <div style="font-size: 11px; font-weight: 600;">✉️ {info.get('email','')}</div>
                </div>
            </header>
            <div class="header-divider"></div>
    """

    text_common = "간기능 | 간염 | 순환기계 | 당뇨 | 췌장기능 | 철결핍성 | 빈혈 | 혈액질환 | 전해질 | 신장기능 | 골격계질환<br>감염성 | 갑상선기능 | 부갑상선기능 | 종양표지자 | 소변 등 80여종 혈액(소변)검사<br>심전도 | 신장 | 체중 | 혈압 | 시력 | 청력 | 체성분 | 건강유형분석 | 폐기능 | 안저 | 안압<br>혈액점도검사 | 유전자20종 | 흉부X-ray | 복부초음파 | 위수면내시경<br>(여)자궁경부세포진 | (여)유방촬영 - #30세이상 권장#"
    text_a = "[01] 갑상선초음파  [10] 골다공증QCT+비타민D\n[02] 경동맥초음파  [11] 혈관협착도ABI\n[03] (여)경질초음파  [12] (여)액상 자궁경부세포진\n[04] 뇌CT  [13] (여) HPV바이러스\n[05] 폐CT  [14] (여)(혈액)마스토체크:유방암\n[06] 요추CT  [15] (혈액)NK뷰키트\n[07] 경추CT  [16] (여)(혈액)여성호르몬\n[08] 심장MDCT  [17] (남)(혈액)남성호르몬\n[09] 복부비만CT"
    text_b = "[가] 대장수면내시경  [마] 부정맥검사S-PATCH\n[나] 심장초음파  [바] [혈액]알레르기검사\n[다] (여)유방초음파  [사] [혈액]알츠온:치매위험도\n[라] [분변]대장암_얼리텍  [아] [혈액]간섬유화검사\n[자] 폐렴예방접종:15가"
    # [수정] 에피클락 삭제됨
    text_c = "[A] 뇌MRI+MRA  [E] [혈액]스마트암검사(남6/여7종)\n[B] 췌장MRI  [F] [혈액]선천적 유전자검사\n[C] 경추MRI\n[D] 요추MRI"

    guide_content = f"""
            <div class="guide-box">
                <span class="guide-title">1. 유동적 그룹 선택 시스템 (Flexible Option)</span>
                <div style="display:flex; justify-content:space-between; align-items: flex-start; gap: 20px;">
                    <div style="flex: 1;">
                        <div style="margin-bottom: 6px; background-color:#ffebee; padding:4px 8px; border-radius:4px; border-left:3px solid #e57373;">
                            • <b>A그룹 2개</b> <span style="color:#aaa">⇄</span> <span class="highlight-text">B그룹 1개</span> 로 변경 선택 가능
                        </div>
                        <div style="margin-bottom: 6px; background-color:#ffebee; padding:4px 8px; border-radius:4px; border-left:3px solid #e57373;">
                            • <b>A그룹 4개</b> <span style="color:#aaa">⇄</span> <span class="highlight-text">C그룹 1개</span> 로 변경 선택 가능
                        </div>
                        <div style="margin-bottom: 6px; padding:2px 5px;">• <span class="highlight-text">유전자검사 20종</span> (기본제공) <span style="color:#aaa">⇄</span> <b>A그룹 1개</b> 로 변경 가능</div>
                        <div style="padding:2px 5px;">• <span class="important-note">공단 위암 대상자</span> 위내시경 진행 시 <span class="highlight-text">A그룹 추가 1가지</span> 선택 가능</div>
                    </div>
                    <div style="flex: 0.8; border-left:3px solid #ddd; padding-left:20px; color:#2c3e50;">
                        <span style="font-weight:bold; display:block; margin-bottom:8px; font-size:13px; color:#c0392b;">[비고: MRI 정밀 장비 안내]</span>
                        <span style="font-weight:bold; font-size:14px; color:#000;">Full Protocol Scan 시행</span><br>
                        <span style="color:#666; font-size:11px;">(진단적 가치 없는 검사는 하지 않습니다.)</span><br>
                        <span class="highlight-text" style="font-size:14px;">최신 3.0T MRI 장비 보유</span>
                    </div>
                </div>
                <div style="margin-top:12px; font-style:italic; color:#666; font-size: 11px; padding-left:5px;">
                * 수면 내시경의 경우 80세까지 진행 가능합니다.
                </div>
            </div>

            <div class="program-grid">
                <div class="grid-box common-box">
                    <div class="grid-header header-common">2. 상세 검진 항목 및 그룹 구성 요약</div>
                    <div class="grid-sub-header">공통 항목 <span style="font-weight:normal;">(위내시경 포함)</span></div>
                    <div class="grid-content">
                        간기능 | 간염 | 순환기계 | 당뇨 | 췌장기능 | 철결핍성 | 빈혈 | 혈액질환 | 전해질 | 신장기능 | 골격계질환<br>
                        감염성 | 갑상선기능 | 부갑상선기능 | 종양표지자 | 소변 등 80여종 혈액(소변)검사<br>
                        심전도 | 신장 | 체중 | 혈압 | 시력 | 청력 | 체성분 | 건강유형분석 | 폐기능 | 안저 | 안압<br>
                        혈액점도검사 | 유전자20종 | 흉부X-ray | 복부초음파 | 위수면내시경<br>
                        (여)자궁경부세포진 | (여)유방촬영 - #30세이상 권장#
                    </div>
                </div>
                <div class="grid-row">
                    <div class="grid-col" style="flex: 1.2;">
                        <div class="grid-box">
                            <div class="grid-header header-a">A 그룹</div>
                            <div class="grid-content-list">
                                <div>[01] 갑상선초음파</div> <div>[10] 골다공증QCT+비타민D</div>
                                <div>[02] 경동맥초음파</div> <div>[11] 혈관협착도ABI</div>
                                <div>[03] (여)경질초음파</div> <div>[12] (여)액상 자궁경부세포진</div>
                                <div>[04] 뇌CT</div> <div>[13] (여) HPV바이러스</div>
                                <div>[05] 폐CT</div> <div>[14] (여)(혈액)마스토체크:유방암</div>
                                <div>[06] 요추CT</div> <div>[15] (혈액)NK뷰키트</div>
                                <div>[07] 경추CT</div> <div>[16] (여)(혈액)여성호르몬</div>
                                <div>[08] 심장MDCT</div> <div>[17] (남)(혈액)남성호르몬</div>
                                <div>[09] 복부비만CT</div>
                            </div>
                        </div>
                    </div>
                    <div class="grid-col" style="flex: 1;">
                        <div class="grid-box">
                            <div class="grid-header header-b">B 그룹</div>
                            <div class="grid-content-list">
                                <div>[가] 대장수면내시경</div> <div>[마] 부정맥검사S-PATCH</div>
                                <div>[나] 심장초음파</div> <div>[바] [혈액]알레르기검사</div>
                                <div>[다] (여)유방초음파</div> <div>[사] [혈액]알츠온:치매위험도</div>
                                <div>[라] [분변]대장암_얼리텍</div> <div>[아] [혈액]간섬유화검사</div>
                                <div></div> <div>[자] 폐렴예방접종:15가</div>
                            </div>
                        </div>
                        <div class="grid-box" style="margin-top:5px; flex-grow:1;">
                            <div class="grid-header header-c">C 그룹</div>
                            <div class="grid-content-list">
                                <div>[A] 뇌MRI+MRA</div> 
                                <div style="letter-spacing:-1.5px; white-space:nowrap;">[E] [혈액]스마트암검사(남6/여7종)</div>
                                <div>[B] 췌장MRI</div> <div>[F] [혈액]선천적 유전자검사</div>
                                <div>[C] 경추MRI</div> <div></div>
                                <div>[D] 요추MRI</div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
    """

    summary_content = f"""
            <div class="section">
                <div class="sec-title">3. 검진 프로그램 요약</div>
                <table class="summary-table">
                    <thead><tr><th style="width:25%">구분</th>{sum_headers}</tr></thead>
                    <tbody>{sum_rows_html}</tbody>
                </table>
            </div>
    """

    table_a = render_table_html("4. A 그룹 ", data.get('A', []))
    table_b = render_table_html("5. B 그룹 ", data.get('B', []), footer="* A그룹 2개를 제외하고 B그룹 1개 선택 가능")
    table_c = render_table_html("6. C 그룹 ", data.get('C', []), footer="* A그룹 4개를 제외하고 C그룹 1개 선택 가능")
    
    equip_data = (data.get('EQUIP', []) or []) + (data.get('COMMON_BLOOD', []) or [])
    
    price_vals = []
    for p in plans:
        txt = p.get('price_txt', p['name'])
        nums = re.findall(r'\d+', str(txt))
        if nums:
            val = int(nums[0]) * 10000
            price_vals.append(f"{val:,}")
        else:
            price_vals.append("-")
            
    equip_data.append({
        "category": "",
        "name": "우대수가",
        "values": price_vals
    })

    table_equip = render_table_html("7. 기초 장비 및 혈액 검사", equip_data, show_sub=True, merge=True, merge_filter=lambda it: bool(it.get("_force_merge_gene")))

    footer = """
            <div style="text-align:center; font-size:11px; color:#7f8c8d; margin-top:30px; padding-top:20px; border-top:1px solid #eee;">
                본 제안서는 귀사의 임직원 건강 증진을 위해 작성되었으며, 세부 검진 항목 및 일정은 협의에 따라 조정될 수 있습니다.
            </div>
        </div>
    </body>
    </html>
    """
    
    return head + cover_html + header_content + guide_content + summary_content + table_a + table_b + table_c + table_equip + footer

def generate_excel_bytes(plans, data, summary, info):
    """엑셀 생성"""
    company = info.get('company', '기업')
    manager_name = info.get('name', '')
    title_text = f"2026 {company} 임직원 건강검진 제안서"
    
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # 1. 표지 시트 생성 - 공문서 스타일
    # ----------------------------------------------------
    ws_cover = wb.active
    ws_cover.title = "표지"
    ws_cover.page_setup.paperSize = 9  # A4
    ws_cover.print_options.horizontalCentered = True
    ws_cover.print_options.verticalCentered = True
    
    thick_bottom = Border(bottom=Side(style='thick', color="1A253A"))
    
    ws_cover['E3'] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"
    ws_cover['E3'].font = Font(size=11, color="555555")
    ws_cover['E3'].alignment = Alignment(horizontal='right')
    ws_cover.merge_cells("E3:H3")

    ws_cover['B15'] = "2026 임직원 건강검진 제안서"
    ws_cover['B15'].font = Font(size=20, bold=True, color="333333")
    ws_cover['B15'].alignment = Alignment(horizontal='center', vertical='center')
    ws_cover.merge_cells("B15:H15")
    
    ws_cover['B17'] = f"{company} 귀중"
    ws_cover['B17'].font = Font(size=36, bold=True, color="1A253A")
    ws_cover['B17'].alignment = Alignment(horizontal='center', vertical='center')
    ws_cover['B17'].border = thick_bottom
    ws_cover.merge_cells("B17:H17")

    start_row = 32
    
    ws_cover[f'E{start_row}'] = "뉴고려병원 검진사업부"
    ws_cover[f'E{start_row}'].font = Font(size=14, bold=True, color="1A253A")
    ws_cover[f'E{start_row}'].alignment = Alignment(horizontal='left')
    ws_cover.merge_cells(f"E{start_row}:H{start_row}")
    
    ws_cover[f'E{start_row+1}'] = f"담당자: {manager_name} 팀장"
    ws_cover[f'E{start_row+1}'].font = Font(size=12, color="333333")
    ws_cover[f'E{start_row+1}'].alignment = Alignment(horizontal='left')
    ws_cover.merge_cells(f"E{start_row+1}:H{start_row+1}")

    ws_cover[f'E{start_row+2}'] = f"T. {info.get('phone','')} / E. {info.get('email','')}"
    ws_cover[f'E{start_row+2}'].font = Font(size=11, color="555555")
    ws_cover[f'E{start_row+2}'].alignment = Alignment(horizontal='left')
    ws_cover.merge_cells(f"E{start_row+2}:H{start_row+2}")

    for r in range(1, 45):
        ws_cover.row_dimensions[r].height = 20
    ws_cover.row_dimensions[17].height = 60

    # ----------------------------------------------------
    # 2. 견적서 상세 시트 생성
    # ----------------------------------------------------
    ws = wb.create_sheet("제안서")
    
    ws.page_setup.paperSize = 9
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.5; ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5; ws.page_margins.bottom = 0.5

    thin_border = Border(left=Side(style='thin',color="CCCCCC"), right=Side(style='thin',color="CCCCCC"), top=Side(style='thin',color="CCCCCC"), bottom=Side(style='thin',color="CCCCCC"))
    box_side = Side(style='medium', color="2C3E50")
    header_fill = PatternFill(start_color="F0F2F5", end_color="F0F2F5", fill_type="solid")
    sum_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def draw_box_border(ws, min_r, max_r, min_c, max_c):
        for c in range(min_c, max_c + 1):
            ws.cell(row=min_r, column=c).border = Border(left=ws.cell(row=min_r, column=c).border.left, right=ws.cell(row=min_r, column=c).border.right, top=box_side, bottom=ws.cell(row=min_r, column=c).border.bottom)
            ws.cell(row=max_r, column=c).border = Border(left=ws.cell(row=max_r, column=c).border.left, right=ws.cell(row=max_r, column=c).border.right, top=ws.cell(row=max_r, column=c).border.top, bottom=box_side)
        for r in range(min_r, max_r + 1):
            ws.cell(row=r, column=min_c).border = Border(left=box_side, right=ws.cell(row=r, column=min_c).border.right, top=ws.cell(row=r, column=min_c).border.top, bottom=ws.cell(row=r, column=min_c).border.bottom)
            ws.cell(row=r, column=max_c).border = Border(left=ws.cell(row=r, column=max_c).border.left, right=box_side, top=ws.cell(row=r, column=max_c).border.top, bottom=ws.cell(row=r, column=max_c).border.bottom)

    # Header
    ws['A1'] = "뉴고려병원"
    ws['A1'].font = Font(size=16, bold=True, color="1A253A")
    ws['A2'] = title_text
    ws['A2'].font = Font(size=14, bold=True)
    ws['A3'] = f"제안일자: {datetime.now().strftime('%Y-%m-%d')}"

    last_col = len(plans) + 1
    if last_col < 3: last_col = 3

    mgr_start_col = last_col - 1 if last_col > 2 else last_col
    ws.merge_cells(start_row=1, start_column=mgr_start_col, end_row=1, end_column=last_col)
    ws.cell(row=1, column=mgr_start_col, value="담당자").font = Font(bold=True, color="7F8C8D")
    ws.cell(row=1, column=mgr_start_col).alignment = Alignment(horizontal='right')
    
    ws.merge_cells(start_row=2, start_column=mgr_start_col, end_row=2, end_column=last_col)
    ws.cell(row=2, column=mgr_start_col, value=f"{info.get('name','')} 팀장").font = Font(bold=True, size=12)
    ws.cell(row=2, column=mgr_start_col).alignment = Alignment(horizontal='right')

    ws.merge_cells(start_row=3, start_column=mgr_start_col, end_row=3, end_column=last_col)
    ws.cell(row=3, column=mgr_start_col, value=info.get('phone','')).alignment = Alignment(horizontal='right')

    ws.merge_cells(start_row=4, start_column=mgr_start_col, end_row=4, end_column=last_col)
    ws.cell(row=4, column=mgr_start_col, value=info.get('email','')).alignment = Alignment(horizontal='right')

    current_row = 6

    # 유동적 그룹
    ws.cell(row=current_row, column=1, value="1. 유동적 그룹 선택 시스템 (Flexible Option)").font = Font(bold=True, size=12, color="2C3E50")
    current_row += 1
    guide_text = (
        "• A그룹 2개 ⇄ B그룹 1개 로 변경 선택 가능\n"
        "• A그룹 4개 ⇄ C그룹 1개 로 변경 선택 가능\n"
        "• 유전자검사 20종 (기본제공) ⇄ A그룹 1개 로 변경 가능\n"
        "• 공단 위암 대상자 위내시경 진행 시 A그룹 추가 1가지 선택 가능\n\n"
        "[비고: MRI 정밀 장비 안내]\n"
        "Full Protocol Scan 시행 (진단적 가치 없는 검사는 하지 않습니다.) / 최신 3.0T MRI 장비 보유\n"
        "수면 내시경의 경우 80세까지 진행 가능합니다."
    )
    start_r = current_row
    end_r = current_row + 6
    ws.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=last_col)
    cell = ws.cell(row=start_r, column=1, value=guide_text)
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
    
    draw_box_border(ws, start_r, end_r, 1, last_col)
    for r in range(start_r, end_r + 1): ws.row_dimensions[r].height = 21 
    current_row += 8

    # 상세 항목
    ws.cell(row=current_row, column=1, value="2. 상세 검진 항목 및 그룹 구성 요약").font = Font(bold=True, size=12, color="2C3E50")
    current_row += 1
    
    # [수정됨] C그룹 텍스트 수정 반영 (에피클락 삭제)
    text_common = "간기능 | 간염 | 순환기계 | 당뇨 | 췌장기능 | 철결핍성 | 빈혈 | 혈액질환 | 전해질 | 신장기능 | 골격계질환\n감염성 | 갑상선기능 | 부갑상선기능 | 종양표지자 | 소변 등 80여종 혈액(소변)검사\n심전도 | 신장 | 체중 | 혈압 | 시력 | 청력 | 체성분 | 건강유형분석 | 폐기능 | 안저 | 안압\n혈액점도검사 | 유전자20종 | 흉부X-ray | 복부초음파 | 위수면내시경\n(여)자궁경부세포진 | (여)유방촬영 - #30세이상 권장#"
    text_a = "[01] 갑상선초음파  [10] 골다공증QCT+비타민D\n[02] 경동맥초음파  [11] 혈관협착도ABI\n[03] (여)경질초음파  [12] (여)액상 자궁경부세포진\n[04] 뇌CT  [13] (여) HPV바이러스\n[05] 폐CT  [14] (여)(혈액)마스토체크:유방암\n[06] 요추CT  [15] (혈액)NK뷰키트\n[07] 경추CT  [16] (여)(혈액)여성호르몬\n[08] 심장MDCT  [17] (남)(혈액)남성호르몬\n[09] 복부비만CT"
    text_b = "[가] 대장수면내시경  [마] 부정맥검사S-PATCH\n[나] 심장초음파  [바] [혈액]알레르기검사\n[다] (여)유방초음파  [사] [혈액]알츠온:치매위험도\n[라] [분변]대장암_얼리텍  [아] [혈액]간섬유화검사\n[자] 폐렴예방접종:15가"
    text_c = "[A] 뇌MRI+MRA  [E] [혈액]스마트암검사(남6/여7종)\n[B] 췌장MRI  [F] [혈액]선천적 유전자검사\n[C] 경추MRI\n[D] 요추MRI"

    box_start_row = current_row
    ws.cell(row=current_row, column=1, value="공통 항목 (위내시경 포함)").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col)
    ws.cell(row=current_row, column=1).alignment = center_align
    current_row += 1
    
    content_start = current_row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+4, end_column=last_col)
    c = ws.cell(row=current_row, column=1, value=text_common)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
    c.border = thin_border
    for r in range(content_start, current_row + 5): ws.row_dimensions[r].height = 20
    draw_box_border(ws, box_start_row, current_row+4, 1, last_col)
    current_row += 5

    def write_group_box(title, text, color_hex, row_h):
        nonlocal current_row
        b_start = current_row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+3, end_column=1)
        cell_h = ws.cell(row=current_row, column=1, value=title)
        cell_h.font = Font(bold=True, color="FFFFFF")
        cell_h.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        cell_h.alignment = center_align

        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+3, end_column=last_col)
        cell_c = ws.cell(row=current_row, column=2, value=text)
        cell_c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left", indent=1)
        cell_c.border = thin_border
        
        for r in range(current_row, current_row + 4): ws.row_dimensions[r].height = row_h
        draw_box_border(ws, b_start, current_row+3, 1, last_col)
        current_row += 4

    write_group_box("A 그룹", text_a, "566573", 39)
    write_group_box("B 그룹", text_b, "7F8C8D", 23)
    write_group_box("C 그룹", text_c, "2C3E50", 15)
    current_row += 1

    # Summary
    ws.cell(row=current_row, column=1, value="3. 검진 프로그램 요약").font = Font(bold=True, size=12)
    current_row += 1
    ws.cell(row=current_row, column=1, value="구분").fill = sum_fill
    ws.cell(row=current_row, column=1).font = white_font
    ws.cell(row=current_row, column=1).alignment = center_align
    for i, p in enumerate(plans):
        c = ws.cell(row=current_row, column=i+2, value=p['name'])
        c.fill = sum_fill; c.font = white_font; c.alignment = center_align
    current_row += 1

    def write_sum_row(title, vals):
        nonlocal current_row
        ws.cell(row=current_row, column=1, value=title).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = thin_border
        ws.cell(row=current_row, column=1).alignment = left_align
        for i, v in enumerate(vals):
            c = ws.cell(row=current_row, column=i+2, value=v)
            c.alignment = center_align; c.border = thin_border
        current_row += 1

    write_sum_row("A그룹", [s['a'] for s in summary])
    write_sum_row("B그룹", [s['b'] for s in summary])
    write_sum_row("C그룹", [s['c'] for s in summary])
    current_row += 1
    
    ws.row_breaks.append(Break(id=current_row))
    current_row += 1

    # 상세
    def write_section(title, items, merge=True, merge_filter=None):
        nonlocal current_row
        if not items: return
        ws.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=12, color="2C3E50")
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="검사 항목").fill = header_fill
        ws.cell(row=current_row, column=1).border = thin_border
        ws.cell(row=current_row, column=1).alignment = center_align
        for i, p in enumerate(plans):
            c = ws.cell(row=current_row, column=i+2, value=p['name'])
            c.fill = header_fill; c.border = thin_border; c.alignment = center_align
        current_row += 1
        
        start_row = current_row
        def norm(v):
            if not v or v in ['-', '미선택', 'X']: return ""
            if '선택' in str(v): return re.sub(r'(선택)\s*(\d+)', r'\1 \2', str(v))
            if 'O' in str(v) or '기본' in str(v): return "O"
            return v

        grid = []
        for item in items:
            row_vals = [norm(v) for v in item['values']]
            grid.append(row_vals)
            name_val = f"[{item['category']}] {item['name']}" if item.get('category') else item['name']
            
            c = ws.cell(row=current_row, column=1, value=name_val)
            c.border = thin_border; c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            if item['name'] == "우대수가":
                c.fill = header_fill
                c.font = Font(bold=True)

            for i, v in enumerate(row_vals):
                c = ws.cell(row=current_row, column=i+2, value=v)
                c.border = thin_border; c.alignment = center_align
                if v == "O": c.font = Font(bold=True)
            current_row += 1
        
        can_merge_row = [merge_filter(it) for it in items] if merge_filter else [True] * len(items)

        if merge:
            for c_idx in range(len(plans)):
                r = 0
                while r < len(grid):
                    val = grid[r][c_idx]
                    if val and can_merge_row[r]:
                        span = 1
                        for k in range(r + 1, len(grid)):
                            if can_merge_row[k] and grid[k][c_idx] == val: span += 1
                            else: break
                        if span > 1:
                            ws.merge_cells(start_row=start_row+r, start_column=c_idx+2, end_row=start_row+r+span-1, end_column=c_idx+2)
                            cell = ws.cell(row=start_row+r, column=c_idx+2)
                            cell.alignment = center_align
                        r += span
                    else: r += 1
        current_row += 2

    write_section("4. A 그룹 ", data['A'])
    write_section("5. B 그룹 ", data['B'])
    write_section("6. C 그룹 ", data['C'])
    
    ws.row_breaks.append(Break(id=current_row))
    current_row += 1
    
    equip_data = (data.get('EQUIP', []) or []) + (data.get('COMMON_BLOOD', []) or [])
    
    price_vals = []
    for p in plans:
        txt = p.get('price_txt', p['name'])
        nums = re.findall(r'\d+', str(txt))
        if nums:
            val = int(nums[0]) * 10000
            price_vals.append(f"{val:,}")
        else:
            price_vals.append("-")
            
    equip_data.append({
        "category": "",
        "name": "우대수가",
        "values": price_vals
    })

    write_section("7. 기초 장비 및 혈액 검사", equip_data, merge=True, merge_filter=lambda it: bool(it.get("_force_merge_gene")))

    ws.column_dimensions['A'].width = 32
    for i in range(len(plans)): ws.column_dimensions[get_column_letter(i+2)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
