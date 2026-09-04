# -*- coding: utf-8 -*-
from __future__ import annotations

import io
import json
import re
from typing import Any, Dict

import openpyxl
import streamlit as st

from flyer_data import normalize_flyer_data, dumps_flyer_json
import flyer_engine as _flyer_engine
from flyer_engine import FlyerEngine, THEMES

# ReportLab은 CFF 방식 OTF를 지원하지 않으므로 TrueType Pretendard를 사용합니다.
_flyer_engine.FONT_URLS.update({
    "regular": "https://raw.githubusercontent.com/wefonts/Pretendard/main/Pretendard-Regular.ttf",
    "medium": "https://raw.githubusercontent.com/wefonts/Pretendard/main/Pretendard-Medium.ttf",
    "semibold": "https://raw.githubusercontent.com/wefonts/Pretendard/main/Pretendard-SemiBold.ttf",
    "bold": "https://raw.githubusercontent.com/wefonts/Pretendard/main/Pretendard-Bold.ttf",
})
try:
    for _cached in _flyer_engine._font_dir().glob("Pretendard-*.otf"):
        _cached.unlink(missing_ok=True)
except Exception:
    pass

HIDDEN_SHEET = "_NK_FLYER_DATA"


def _lines(value):
    if isinstance(value, list):
        return "\n".join(str(x) for x in value)
    return str(value or "")


def _parse_lines(value):
    return [x.strip() for x in str(value or "").splitlines() if x.strip()]


def embed_quote_payload(html: str, flyer_data: Dict[str, Any]) -> str:
    """견적서 HTML 안에 안내문 재생성용 데이터를 저장합니다."""
    payload = json.dumps(normalize_flyer_data(flyer_data), ensure_ascii=False, separators=(",", ":"))
    payload = payload.replace("</", "<\\/")
    block = f'\n<script id="nk-flyer-data" type="application/json">{payload}</script>\n'
    if re.search(r"</body\s*>", html, re.I):
        return re.sub(r"</body\s*>", block + "</body>", html, count=1, flags=re.I)
    return html + block


def extract_quote_payload(html: str):
    m = re.search(r'<script[^>]+id=["\']nk-flyer-data["\'][^>]*>([\s\S]*?)</script>', html, re.I)
    if not m:
        return None
    raw = m.group(1).replace("<\\/", "</")
    try:
        value = json.loads(raw)
    except Exception:
        return None
    return normalize_flyer_data(value) if isinstance(value, dict) else None


def embed_flyer_in_excel(excel_bytes: bytes, flyer_data: Dict[str, Any]) -> bytes:
    """다운로드 견적서 XLSX에 안내문 데이터를 숨김 시트로 저장합니다."""
    src = io.BytesIO(excel_bytes)
    wb = openpyxl.load_workbook(src)
    if HIDDEN_SHEET in wb.sheetnames:
        del wb[HIDDEN_SHEET]
    ws = wb.create_sheet(HIDDEN_SHEET)
    ws["A1"] = json.dumps(normalize_flyer_data(flyer_data), ensure_ascii=False, separators=(",", ":"))
    ws.sheet_state = "veryHidden"
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


def extract_flyer_from_excel(excel_bytes: bytes):
    """이 앱에서 생성한 XLSX의 숨김 시트에서 안내문 데이터를 복원합니다."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=False, data_only=False)
        if HIDDEN_SHEET not in wb.sheetnames:
            wb.close()
            return None
        raw = wb[HIDDEN_SHEET]["A1"].value
        wb.close()
        value = json.loads(str(raw or ""))
        return normalize_flyer_data(value) if isinstance(value, dict) else None
    except Exception:
        return None


def render_flyer_editor(initial: Dict[str, Any], prefix: str = "flyer", allow_import: bool = False):
    """안내문 편집 + PDF/PNG 생성 UI. 반환값은 현재 데이터."""
    data = normalize_flyer_data(initial)
    st.subheader("건강검진 안내문")
    st.caption("견적서에서 설정한 플랜명·금액·A/B/C 선택 규칙이 자동 반영됩니다. 아래에서 안내문 전용 내용만 추가 수정할 수 있습니다.")

    c1, c2, c3 = st.columns(3)
    with c1:
        data["title"] = st.text_input("안내문 제목", value=data.get("title", ""), key=f"{prefix}_title")
        data["target"] = st.text_input("검진 대상", value=data.get("target", ""), key=f"{prefix}_target")
    with c2:
        data["period"] = st.text_input("검진기간", value=data.get("period", ""), key=f"{prefix}_period")
        data["application_period"] = st.text_input("접수기간", value=data.get("application_period", ""), key=f"{prefix}_app_period")
    with c3:
        data["phone"] = st.text_input("검진 문의", value=data.get("phone", "1833-9988"), key=f"{prefix}_phone")
        data["qr_url"] = st.text_input("QR 연결 주소", value=data.get("qr_url", ""), key=f"{prefix}_qr")

    c4, c5 = st.columns([1, 2])
    with c4:
        themes = list(THEMES.keys())
        cur_theme = data.get("theme", "기업형")
        idx = themes.index(cur_theme) if cur_theme in themes else 0
        data["theme"] = st.selectbox("안내문 테마", themes, index=idx, key=f"{prefix}_theme")
        bg = st.file_uploader("배경 이미지 PNG/JPG (선택)", type=["png", "jpg", "jpeg"], key=f"{prefix}_bg")
    with c5:
        data["event_title"] = st.text_input("이벤트/혜택 제목", value=data.get("event_title", ""), key=f"{prefix}_event_title")
        data["event_lines"] = _parse_lines(st.text_area("혜택 문구 (한 줄당 1개)", value=_lines(data.get("event_lines", [])), height=85, key=f"{prefix}_event_lines"))

    st.markdown("#### 견적서 플랜 → 안내문 자동 반영")
    packages = data.get("packages", [])
    if not packages:
        st.info("아직 반영할 견적 플랜이 없습니다.")
    else:
        for i, p in enumerate(packages):
            cols = st.columns([1.5, 5, 1.2])
            p["name"] = cols[0].text_input("플랜명", value=p.get("name", ""), key=f"{prefix}_pkg_name_{i}")
            p["detail"] = cols[1].text_input("구성", value=p.get("detail", ""), key=f"{prefix}_pkg_detail_{i}")
            p["price"] = cols[2].text_input("금액", value=p.get("price") or p.get("male_price", ""), key=f"{prefix}_pkg_price_{i}")
            p["male_price"] = p["price"]
            p["female_price"] = p["price"]
        data["packages"] = packages

    with st.expander("공통항목 / A·B·C 그룹 수정", expanded=False):
        data["common_items"] = st.text_area("공통항목", value=data.get("common_items", ""), height=110, key=f"{prefix}_common")
        a, b, c = st.columns(3)
        data["groups"]["A"] = _parse_lines(a.text_area("A그룹", value=_lines(data.get("groups", {}).get("A", [])), height=330, key=f"{prefix}_ga"))
        data["groups"]["B"] = _parse_lines(b.text_area("B그룹", value=_lines(data.get("groups", {}).get("B", [])), height=330, key=f"{prefix}_gb"))
        data["groups"]["C"] = _parse_lines(c.text_area("C그룹", value=_lines(data.get("groups", {}).get("C", [])), height=330, key=f"{prefix}_gc"))

    data["notes"] = _parse_lines(st.text_area("하단 안내", value=_lines(data.get("notes", [])), height=70, key=f"{prefix}_notes"))

    if st.button("안내문 PDF / PNG 생성", type="primary", key=f"{prefix}_generate"):
        try:
            engine = FlyerEngine()
            background = bg.getvalue() if bg else None
            st.session_state[f"{prefix}_pdf"] = engine.render_pdf(data, background)
            st.session_state[f"{prefix}_png"] = engine.render_png(data, background)
            st.session_state[f"{prefix}_json"] = dumps_flyer_json(data)
            st.success("안내문 생성 완료")
        except Exception as e:
            st.error(f"안내문 생성 실패: {e}")

    if st.session_state.get(f"{prefix}_png"):
        st.image(st.session_state[f"{prefix}_png"], use_container_width=True)
        d1, d2, d3 = st.columns(3)
        d1.download_button("📥 안내문 PDF", st.session_state[f"{prefix}_pdf"], "건강검진_안내문.pdf", "application/pdf", key=f"{prefix}_dl_pdf")
        d2.download_button("📥 안내문 PNG", st.session_state[f"{prefix}_png"], "건강검진_안내문.png", "image/png", key=f"{prefix}_dl_png")
        d3.download_button("📥 안내문 데이터 JSON", st.session_state[f"{prefix}_json"], "flyer_data.json", "application/json", key=f"{prefix}_dl_json")

    return normalize_flyer_data(data)
